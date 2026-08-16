# -*- coding: utf-8 -*-
"""t0 시뮬레이션(GIF) — AVT animate_results.animate_jump 재사용 (사용자 표준 스틱피겨+4절링크+2×2 상태창).

경로: xlsx를 animate_results 스키마(time/base_height/q_1/q_2/dz/dq_1/dq_2/tau_1/tau_2/grf_z
[+q_m/l_i])로 쓰고, AVT 폴더의 animate_jump()를 그대로 호출.
"""
import sys
from pathlib import Path

import numpy as np
import openpyxl

HERE = Path(__file__).parent
AVT = Path(r"C:\Users\junho\CVT\AVT LEG\optimization_tasks")
sys.path.insert(0, str(AVT))
sys.path.insert(0, str(HERE))

from t0_figs import _chan, _log_chan   # 채널 규약 재사용


def write_xlsx(ch, out_xlsx, qm=None, l_i_mm=None):
    """채널 dict → animate_results 스키마 xlsx. qm/l_i_mm 주면 CVT 4절링크 실측 표시."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "jump_results_basic"
    heads = ["time", "base_height", "q_1", "q_2", "dz", "dq_1", "dq_2",
             "tau_1", "tau_2", "grf_z"]
    cols = [ch["t"], ch["bz"], ch["q1"], ch["q2"], ch["dz"], ch["dq1"], ch["dq2"],
            ch["tau1"], ch["tau2"], ch["grf"]]
    if qm is not None:
        heads += ["q_m", "l_i"]
        cols += [qm, np.full(len(ch["t"]), l_i_mm)]
    for j, h in enumerate(heads, 1):
        ws.cell(row=1, column=j, value=h)
    for i in range(len(ch["t"])):
        for j, arr in enumerate(cols, 1):
            ws.cell(row=i + 2, column=j, value=float(arr[i]))
    wb.save(out_xlsx)
    return out_xlsx


def gif_of(ch, out_gif, title, qm=None, l_i_mm=None, decim=10):
    """채널 → xlsx → animate_jump GIF. decim: 0.5ms 로그를 5ms로 솎아 프레임 수 제어."""
    if decim > 1:
        ch = {k: np.asarray(v)[::decim] for k, v in ch.items()}
        if qm is not None:
            qm = np.asarray(qm)[::decim]
    xlsx = Path(out_gif).with_suffix(".xlsx")
    write_xlsx(ch, xlsx, qm=qm, l_i_mm=l_i_mm)
    from animate_results import animate_jump
    animate_jump(str(xlsx), save_path=str(out_gif), title=title)
    return out_gif


def gif_of_npz(npz_path, out_gif, title, decim=10):
    z = np.load(npz_path)
    ch = _chan(z)
    qm = None
    li = None
    if "qm" in z.files:
        t = np.asarray(z["t"], float)
        qm = np.asarray(z["qm"], float)[t >= 0]
        li = float(np.atleast_1d(z["l_i"])[0])
    return gif_of(ch, out_gif, title, qm=qm, l_i_mm=li, decim=decim)
