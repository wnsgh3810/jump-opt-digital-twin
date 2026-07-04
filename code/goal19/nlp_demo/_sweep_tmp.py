import casadi as ca
import numpy as np
JF_HIP  = 0.488
JF_KNEE = 0.0141
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.gridspec import GridSpec
from scipy.interpolate import interp1d
import openpyxl
import os
import sys
import builtins

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")

# ============================================================
# Contact Model 선택 가이드
# ============================================================
CONTACT_MODEL = 'hard'   # 'hard' | 'alpha' | 'soft' | 'soft_alpha'
ALPHA         = 0.85    # soft_alpha: body가 느끼는 GRF 비율 (alpha 불필요)
K_C           = 5000   # [N/m]   접촉 스프링 강성
B_C           = 20     # [N·s/m] 접촉 감쇠계수
RAIL_FRICTION = 0.0    # [N·s/m] 레일 점성 마찰
JOINT_FRICTION = 0.0   # [Nm·s/rad] 관절 점성 마찰
# ★ 최종 파라미터 (3600개 sweep, 입력토크sat + z_kin높이 + 입력에너지, score=6):
#   alpha=1.0 (불필요), k_c=3000, b_c=80, tau_lim=15
#   rail_friction=0, joint_friction=0.2
#   vs Real P40: dq2=24.8(0.8%), Imp=20.3(0.5%), E=36.1(2.2%)
#   E/(Imp)^2 = 1.008 (Real: 1.026), h=0.787(0.9%)

def integrate_trapezoid(y, x):
    y = np.asarray(y, dtype=float)
    x = np.asarray(x, dtype=float)
    valid = np.isfinite(y) & np.isfinite(x)
    y = y[valid]
    x = x[valid]
    if len(y) < 2:
        return np.nan
    if hasattr(np, "trapezoid"):
        return np.trapezoid(y, x)
    elif hasattr(np, "trapz"):
        return np.trapz(y, x)
    else:
        raise AttributeError("numpy에 trapezoid/trapz가 모두 없습니다.")


def summarize_peak_event(name, idx, t_arr, dq_arr, tau_arr, grf_z_arr, speed_limit_arr, power_arr):
    dq_val = float(dq_arr[idx])
    tau_val = float(tau_arr[idx])
    grf_val = float(grf_z_arr[idx])
    lim_val = float(speed_limit_arr[idx])
    sat_pct = 100.0 * abs(dq_val) / lim_val if lim_val > 1e-9 else np.nan
    return {
        'name': name,
        'idx': int(idx),
        'time': float(t_arr[idx]),
        'dq': dq_val,
        'tau': tau_val,
        'grf_z': grf_val,
        'speed_limit': lim_val,
        'sat_pct': sat_pct,
        'power': float(power_arr[idx]),
    }


def clean_console_text(text):
    text = str(text)
    if "Impulse_z" in text and ":" in text:
        suffix = text.split(":", 1)[1]
        suffix = suffix.replace("N·s", "N*s").replace("N쨌s", "N*s").replace("N﹞s", "N*s")
        return f"  Impulse_z (Integral Fz dt):{suffix}"
    text = text.replace("← optimizer target", "<- optimizer target")
    text = text.replace("??optimizer target", "<- optimizer target")
    text = text.replace("∠ optimizer target", "<- optimizer target")
    text = text.replace("N쨌s", "N*s")
    text = text.replace("N﹞s", "N*s")
    text = text.replace("N·s", "N*s")
    text = text.replace("(⊕Fz dt)", "(Integral Fz dt)")
    text = text.replace("(◎Fz dt)", "(Integral Fz dt)")
    stripped = text.strip()
    if stripped and len(stripped) >= 20:
        uniq = set(stripped)
        if uniq.issubset({"?", "式", "─"}):
            return "-" * len(stripped)
        if not any(ch.isalnum() for ch in stripped):
            return "-" * len(stripped)
    return text


def clean_print(*args, **kwargs):
    cleaned_args = [clean_console_text(arg) for arg in args]
    return builtins.print(*cleaned_args, **kwargs)


def optimize_jumping(contact_model=CONTACT_MODEL, alpha=ALPHA, k_c=K_C, b_c=B_C):
    print = clean_print
    print(f"\n{'='*60}")
    print(f"  Contact Model : {contact_model.upper()}", end="")
    if contact_model == 'alpha':
        print(f"  (alpha={alpha:.4f})", end="")
    elif contact_model in ('soft', 'soft_alpha'):
        print(f"  (k_c={k_c:.0f} N/m, b_c={b_c:.1f} N/s/m)", end="")
    elif contact_model == 'soft_alpha':
        print(f"  (alpha={alpha:.4f}, k_c={k_c:.0f}, b_c={b_c:.1f})", end="")
    print(f"\n{'='*60}")

    # === 1. System Parameters ===
    params = {
        'M':  1.27424,
        'm1': 1.00537,
        'm2': 0.30422,
        'm_c': 0.54081,
        'm_p': 0.23470,
        'l1': 0.25,
        'l2': 0.25,
        'l_c': 0.03,
        'r1': 0.11526,
        'r2': 0.13909,
        'r_c': 0.02069,
        'r_p': 0.13258,
        'I1': 0.0063422,
        'I2': 0.0035720,
        'I_c': 0.0005797,
        'I_p': 0.0008858,
        'arm_knee': 0.0035,
        'g':  9.81,
        'mu': 0.3,
        'l_o': 0.03,
        'alpha': alpha,
        'k_c':   k_c,
        'b_c':   b_c,
    }

    N = 60
    # Explicit z_kin bound REMOVED. Bounds derived from inverse kinematics
    # (foot_x=0, l1=l2=0.25): z_kin = 0.5*cos(q2/2) and q2 = -pi - 2*q1.
    # User-specified q1 in [-72, -17] deg maps to q2 in [-146, -36] deg,
    # which gives z_kin in [0.146, 0.476] m (rail hardware range ~[0.15, 0.475]).
    q1_lb, q1_ub = -1.2566, -0.2967  # [-72.00 deg, -17.00 deg]
    q2_lb, q2_ub = -2.5482, -0.6283  # [-146.00 deg, -36.00 deg]
    tau_lim = 18.0   # AK80-9 V2 peak (G20)
    dq_lim  = 50.0
    speed_torque_coeff  = -0.731019
    speed_torque_offset =  48.476878

    M_tot_val = params['M'] + params['m1'] + params['m2'] + params['m_c'] + params['m_p']

    # === 2. Optimization Variables ===
    opti = ca.Opti()
    T_st = opti.variable()
    dt   = T_st / (N - 1)

    X    = opti.variable(3, N)   # [z, q1, q2]
    V    = opti.variable(3, N)   # [dz, dq1, dq2]
    U_tau = opti.variable(2, N)  # [tau1, tau2]

    z_pos = X[0, :];  q1 = X[1, :];  q2 = X[2, :]
    dz    = V[0, :];  dq1 = V[1, :]; dq2 = V[2, :]

    if contact_model in ('hard', 'alpha'):
        U_grf = opti.variable(2, N)
        d_x   = U_grf[0, :]
        d_z   = U_grf[1, :]
    else:  # soft
        U_grf_x = opti.variable(1, N)
        d_x     = U_grf_x[0, :]

    # ── Soft Contact Helper ──────────────────────────────────────
    def soft_grf(z_k, q1_k, q2_k, dz_k, dq1_k, dq2_k):
        l1, l2 = params['l1'], params['l2']
        foot_z  = z_k + l1*ca.sin(q1_k) + l2*ca.sin(q1_k + q2_k)
        foot_vz = (dz_k
                   + l1*ca.cos(q1_k)*dq1_k
                   + l2*ca.cos(q1_k + q2_k)*(dq1_k + dq2_k))
        delta     = -foot_z        # 양수 = 발이 지면 아래
        delta_dot = -foot_vz       # 양수 = 침투 속도 증가
        grf_z = params['k_c'] * delta + params['b_c'] * delta_dot
        return grf_z, delta, delta_dot

    # === 3. Dynamics Function ===
    def get_dynamics(x_s, v_s, u_tau, u_grf):
        z, th1, th2     = x_s[0], x_s[1], x_s[2]
        dz_v, dth1, dth2 = v_s[0], v_s[1], v_s[2]
        tau1, tau2      = u_tau[0], u_tau[1]
        dx, dz_grf      = u_grf[0], u_grf[1]
        p = params

        s1,  c1  = ca.sin(th1), ca.cos(th1)
        s2,  c2  = ca.sin(th2), ca.cos(th2)
        s12, c12 = ca.sin(th1+th2), ca.cos(th1+th2)

        M_tot = p['M']+p['m1']+p['m_c']+p['m_p']+p['m2']
        A = p['m1']*p['r1'] + p['m_p']*p['r_p'] + p['m2']*p['l1']
        B = p['m2']*p['r2'] - p['m_c']*p['r_c'] - p['m_p']*p['l_c']
        K = p['m2']*p['l1']*p['r2'] - p['m_p']*p['l_c']*p['r_p']
        I_sig1 = ((p['I1']+p['m1']*p['r1']**2)
                  +(p['I_c']+p['m_c']*p['r_c']**2)
                  +(p['I_p']+p['m_p']*p['r_p']**2+p['m_p']*p['l_c']**2)
                  +(p['I2']+p['m2']*p['r2']**2+p['m2']*p['l1']**2))
        I_sig2 = ((p['I2']+p['m2']*p['r2']**2)
                  +(p['I_c']+p['m_c']*p['r_c']**2)
                  +p['m_p']*p['l_c']**2
                  +p.get('arm_knee', 0.0))

        M_mat = ca.vertcat(
            ca.horzcat(M_tot,           A*c1+B*c12,              B*c12   ),
            ca.horzcat(A*c1+B*c12,      I_sig1+2*K*c2,           I_sig2+K*c2),
            ca.horzcat(B*c12,           I_sig2+K*c2,             I_sig2  ),
        )

        C_vec = ca.vertcat(
            -A*s1*dth1**2 - B*s12*(dth1+dth2)**2,
            -K*s2*(2*dth1*dth2 + dth2**2),
             K*s2*dth1**2,
        )
        G_vec = ca.vertcat(
            M_tot*p['g'],
            p['g']*(A*c1 + B*c12),
            p['g']*B*c12,
        )

        # ALPHA FIX (2026-05-26): symmetric alpha on (dx, dz_grf) — contact compliance
        # affects normal AND tangential consistently. Friction cone |dx| <= mu*dz_grf
        # remains valid (both sides scale equally).
        if contact_model in ('alpha', 'soft_alpha'):
            eff_grf_z = p['alpha'] * dz_grf
            eff_grf_x = p['alpha'] * dx
        else:
            eff_grf_z = dz_grf
            eff_grf_x = dx

        RHS_vec = ca.vertcat(
            eff_grf_z,
            tau1 - eff_grf_x*(p['l1']*s1+p['l2']*s12) + eff_grf_z*(p['l1']*c1+p['l2']*c12),
            tau2 - eff_grf_x*(p['l2']*s12)             + eff_grf_z*(p['l2']*c12),
        )

        F_friction = ca.vertcat(
            -RAIL_FRICTION * dz_v,
            -JF_HIP * dth1,
            -JF_KNEE * dth2,
        )

        return ca.solve(M_mat, RHS_vec - C_vec - G_vec + F_friction)

    # === 4. Direct Collocation ===
    for k in range(N - 1):
        x_k  = X[:, k];   v_k  = V[:, k];   tau_k  = U_tau[:, k]
        x_n  = X[:, k+1]; v_n  = V[:, k+1]; tau_n  = U_tau[:, k+1]

        if contact_model in ('soft', 'soft_alpha'):
            gz_k, _, _ = soft_grf(X[0,k], X[1,k], X[2,k], V[0,k], V[1,k], V[2,k])
            gz_n, _, _ = soft_grf(X[0,k+1], X[1,k+1], X[2,k+1], V[0,k+1], V[1,k+1], V[2,k+1])
            grf_k = ca.vertcat(U_grf_x[0, k],   gz_k)
            grf_n = ca.vertcat(U_grf_x[0, k+1], gz_n)
        else:
            grf_k = U_grf[:, k]
            grf_n = U_grf[:, k+1]

        acc_k = get_dynamics(x_k, v_k, tau_k, grf_k)
        acc_n = get_dynamics(x_n, v_n, tau_n, grf_n)

        opti.subject_to(x_n == x_k + 0.5*dt*(v_k + v_n))
        opti.subject_to(v_n == v_k + 0.5*dt*(acc_k + acc_n))

    # === 5. Kinematic / Contact Constraints ===
    for k in range(N):
        l1, l2 = params['l1'], params['l2']
        rel_x = l1*ca.cos(q1[k]) + l2*ca.cos(q1[k]+q2[k])
        rel_y = l1*ca.sin(q1[k]) + l2*ca.sin(q1[k]+q2[k])
        mu    = params['mu']

        if contact_model in ('hard', 'alpha'):
            opti.subject_to(z_pos[k] + rel_y == 0)
            opti.subject_to(rel_x == 0)
            opti.subject_to(opti.bounded(-mu*d_z[k], d_x[k], mu*d_z[k]))

        else:  # soft
            opti.subject_to(rel_x == 0)
            gz_k, delta_k, _ = soft_grf(z_pos[k], q1[k], q2[k], dz[k], dq1[k], dq2[k])
            opti.subject_to(delta_k >= 0)
            opti.subject_to(gz_k >= 0)
            opti.subject_to(opti.bounded(-mu*gz_k, d_x[k], mu*gz_k))

    # === 6. Physical / Actuator Constraints ===
    opti.subject_to(opti.bounded(0.05, T_st, 0.3))
    # NOTE: Explicit z_kin bound REMOVED. z_kin = -(l1*sin(q1) + l2*sin(q1+q2))
    # is now implicitly limited by the tightened (q1, q2) angle bounds below.
    opti.subject_to(opti.bounded(q1_lb, q1, q1_ub))
    opti.subject_to(opti.bounded(q2_lb, q2, q2_ub))
    opti.subject_to(opti.bounded(-dq_lim, V, dq_lim))
    # 입력 토크 기준 saturation: τ_input = τ_output + friction
    for k in range(N):
        tau1_input_k = U_tau[0, k] + JF_HIP * V[1, k]
        tau2_input_k = U_tau[1, k] + JF_KNEE * V[2, k]
        opti.subject_to(opti.bounded(-tau_lim, tau1_input_k, tau_lim))
        opti.subject_to(opti.bounded(-tau_lim, tau2_input_k, tau_lim))

    if contact_model in ('hard', 'alpha'):
        opti.subject_to(d_z >= 0)
        opti.subject_to(U_grf[1, -1] == 0)

    for k in range(N):
        for j in range(2):
            lim = speed_torque_coeff*ca.fabs(U_tau[j,k]) + speed_torque_offset
            opti.subject_to(ca.fabs(V[j+1, k]) <= lim)

    opti.subject_to(dz[0]  == 0)
    opti.subject_to(dq1[0] == 0)
    opti.subject_to(dq2[0] == 0)

    if contact_model in ('hard', 'alpha'):
        opti.subject_to(U_grf[1, 0] == M_tot_val * params['g'])
    else:
        delta_static = M_tot_val * params['g'] / params['k_c']
        foot_z0 = z_pos[0] + params['l1']*ca.sin(q1[0]) + params['l2']*ca.sin(q1[0]+q2[0])
        opti.subject_to(foot_z0 == -delta_static)
        gz_last, delta_last, ddot_last = soft_grf(
            z_pos[-1], q1[-1], q2[-1], dz[-1], dq1[-1], dq2[-1])
        opti.subject_to(gz_last == 0)

    # === 7. Objective ===
    z_to = z_pos[-1]
    A_expr = params['m1']*params['r1'] + params['m_p']*params['r_p'] + params['m2']*params['l1']
    B_expr = params['m2']*params['r2'] - params['m_c']*params['r_c'] - params['m_p']*params['l_c']

    v_com_z_end = (dz[-1]
                   + (A_expr*ca.cos(q1[-1])*dq1[-1]
                      + B_expr*ca.cos(q1[-1]+q2[-1])*(dq1[-1]+dq2[-1])) / M_tot_val)
    h_base_via_com = z_to + v_com_z_end**2 / (2*params['g'])

    if contact_model in ('hard', 'alpha'):
        J_smooth = sum(
            ca.sumsqr(U_tau[:, k+1] - U_tau[:, k]) * 10
            + ca.sumsqr(U_grf[:, k+1] - U_grf[:, k])
            for k in range(N-1))
    else:
        J_smooth = sum(
            ca.sumsqr(U_tau[:, k+1] - U_tau[:, k]) * 10
            + (U_grf_x[0, k+1] - U_grf_x[0, k])**2
            for k in range(N-1))

    J_smooth_v2 = sum(ca.sumsqr(V[:, k+1] - 2*V[:, k] + V[:, k-1]) for k in range(1, N-1))
    opti.minimize(-2000.0 * h_base_via_com + 0.1 * J_smooth + 20.0 * J_smooth_v2)

    # === 8. Initial Guess ===
    opti.set_initial(T_st, 0.14)
    # Initial guess: crouch (z_kin~0.15) -> takeoff (z_kin~0.475). New tight bounds
    # require initial guesses to stay inside [q_lb[1:], q_ub[1:]].
    opti.set_initial(q1, np.linspace(-0.32, -1.25, N))     # crouch q1 -> takeoff q1
    opti.set_initial(q2, np.linspace(-2.50, -0.65, N))     # crouch q2 -> takeoff q2
    opti.set_initial(U_tau, 9.0 * np.ones((2, N)))

    if contact_model in ('hard', 'alpha'):
        opti.set_initial(d_z, np.full(N, M_tot_val * params['g'] * 2))
    else:
        delta_s = M_tot_val * params['g'] / params['k_c']
        z_guess = np.linspace(0.30 + delta_s, 0.45 + delta_s, N)
        opti.set_initial(z_pos, z_guess)

    # === 9. Solve ===
    opts = {
        'ipopt.print_level': 5,
        'ipopt.max_iter': 8000,
        'ipopt.tol': 1e-4,
        'ipopt.warm_start_init_point': 'yes',
    }
    opti.solver('ipopt', opts)

    try:
        sol = opti.solve()
        print("\nSUCCESS: Optimization Converged")
    except Exception as e:
        print(f"\nFAIL: {e}")
        sol = opti.debug

    # === 10. Extract Results ===
    T_val     = sol.value(T_st)
    t_val     = np.linspace(0, T_val, N)
    x_val     = sol.value(X)
    v_val     = sol.value(V)
    u_tau_val = sol.value(U_tau)

    if contact_model in ('hard', 'alpha'):
        u_grf_val = sol.value(U_grf)
        # ALPHA FIX: saved dz_grf stays RAW.
        if False:
            u_grf_val[1, :] = params['alpha'] * u_grf_val[1, :]
    else:
        u_grf_x_val = sol.value(U_grf_x)
        grf_z_arr   = np.zeros(N)
        delta_arr   = np.zeros(N)
        l1, l2 = params['l1'], params['l2']
        for k in range(N):
            fz = x_val[0,k] + l1*np.sin(x_val[1,k]) + l2*np.sin(x_val[1,k]+x_val[2,k])
            fvz = (v_val[0,k]
                   + l1*np.cos(x_val[1,k])*v_val[1,k]
                   + l2*np.cos(x_val[1,k]+x_val[2,k])*(v_val[1,k]+v_val[2,k]))
            delta_arr[k] = -fz
            grf_z_arr[k] = params['k_c']*(-fz) + params['b_c']*(-fvz)
        u_grf_val = np.vstack([u_grf_x_val, grf_z_arr])

    # Power & Energy 계산 (엑셀 저장을 위해 위로 끌어올림)
    P_hip   = u_tau_val[0,:] * v_val[1,:]
    P_knee  = u_tau_val[1,:] * v_val[2,:]
    P_total = P_hip + P_knee
    
    hip_power_abs = np.abs(P_hip)
    knee_power_abs = np.abs(P_knee)
    total_power_abs = np.abs(P_total)

    # === 11. Save Results to Excel (Interpolated) ===
    dt_save = 0.002
    T_total = T_val
    t_interp = np.arange(0, T_total, dt_save)

    # base_height = z_kin (발끝 z=0 기준, 하드웨어 기준)
    z_kin_arr = -(params['l1']*np.sin(x_val[1,:]) + params['l2']*np.sin(x_val[1,:]+x_val[2,:]))
    raw_data = {
        'base_height': z_kin_arr,
        'q_1': x_val[1, :],
        'q_2': x_val[2, :], 
        'dz': v_val[0, :],
        'dq_1': v_val[1, :],
        'dq_2': v_val[2, :],
        'tau_1': u_tau_val[0, :],
        'tau_2': u_tau_val[1, :],
        'grf_x': u_grf_val[0, :],
        'grf_z': u_grf_val[1, :],
        'hip_power': P_hip,
        'knee_power': P_knee,
        'total_power': P_total,
        'hip_power_abs': hip_power_abs,
        'knee_power_abs': knee_power_abs,
        'total_power_abs': total_power_abs,
    }

    interp_data = {}
    for key, vals in raw_data.items():
        f = interp1d(t_val, vals, kind='linear', fill_value='extrapolate')
        interp_data[key] = f(t_interp)

    # Final-sample correction ?? derived channels? ?? ??? ??.
    tau1_input_arr = u_tau_val[0,:] + JF_HIP * v_val[1,:]
    tau2_input_arr = u_tau_val[1,:] + JF_KNEE * v_val[2,:]
    grf_z_spring_arr = u_grf_val[1,:]
    alpha_val = params.get('alpha', 1.0) if contact_model in ('alpha', 'soft_alpha') else 1.0
    grf_z_body_arr = grf_z_spring_arr * alpha_val
    pre_extra_sources = {
        'tau_1_input': tau1_input_arr,
        'tau_2_input': tau2_input_arr,
        'grf_z_spring': grf_z_spring_arr,
        'grf_z_body': grf_z_body_arr,
    }
    for key, data_src in pre_extra_sources.items():
        f = interp1d(t_val, data_src, kind='linear', fill_value='extrapolate')
        interp_data[key] = f(t_interp)

    # 마지막 행 보정 (에러 방지)
    interp_data['base_height'][-1] = z_kin_arr[-1]
    interp_data['q_1'][-1] = x_val[1, -1]
    interp_data['q_2'][-1] = x_val[2, -1]
    interp_data['dz'][-1] = v_val[0, -1]
    interp_data['dq_1'][-1] = v_val[1, -1]
    interp_data['dq_2'][-1] = v_val[2, -1]
    interp_data['tau_1'][-1] = u_tau_val[0, -1]
    interp_data['tau_2'][-1] = u_tau_val[1, -1]
    interp_data['grf_x'][-1] = u_grf_val[0, -1]
    interp_data['grf_z'][-1] = u_grf_val[1, -1]
    interp_data['grf_z_spring'][-1] = u_grf_val[1, -1]
    alpha_val = params.get('alpha', 1.0) if contact_model in ('alpha', 'soft_alpha') else 1.0
    interp_data['grf_z_body'][-1] = u_grf_val[1, -1] * alpha_val
    interp_data['tau_1_input'][-1] = u_tau_val[0, -1] + JF_HIP * v_val[1, -1]
    interp_data['tau_2_input'][-1] = u_tau_val[1, -1] + JF_KNEE * v_val[2, -1]
    interp_data['hip_power'][-1] = P_hip[-1]
    interp_data['knee_power'][-1] = P_knee[-1]
    interp_data['total_power'][-1] = P_total[-1]
    interp_data['hip_power_abs'][-1] = hip_power_abs[-1]
    interp_data['knee_power_abs'][-1] = knee_power_abs[-1]
    interp_data['total_power_abs'][-1] = total_power_abs[-1]

    N_interp = len(t_interp)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "jump_results_basic"

    # 입력 토크 = 출력 토크 + 마찰 보상
    tau1_input_arr = u_tau_val[0,:] + JF_HIP * v_val[1,:]
    tau2_input_arr = u_tau_val[1,:] + JF_KNEE * v_val[2,:]
    # Alpha 전후 GRF
    grf_z_spring_arr = u_grf_val[1,:]
    alpha_val = params.get('alpha', 1.0) if contact_model in ('alpha', 'soft_alpha') else 1.0
    grf_z_body_arr = grf_z_spring_arr * alpha_val

    # 보간에 입력토크, alpha GRF 추가
    extra_sources = {
        'tau_1_input': (t_val, tau1_input_arr),
        'tau_2_input': (t_val, tau2_input_arr),
        'grf_z_spring': (t_val, grf_z_spring_arr),
        'grf_z_body': (t_val, grf_z_body_arr),
    }
    for key, (t_src, data_src) in extra_sources.items():
        f = interp1d(t_src, data_src, kind='linear', fill_value='extrapolate')
        interp_data[key] = f(t_interp)

    headers = [
        'time', 'base_height', 'q_1', 'q_2', 'dz', 'dq_1', 'dq_2',
        'tau_1_output', 'tau_2_output', 'tau_1_input', 'tau_2_input',
        'grf_x', 'grf_z_spring', 'grf_z_body',
        'hip_power', 'knee_power', 'total_power',
        'hip_power_abs', 'knee_power_abs', 'total_power_abs'
    ]

    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for k in range(N_interp):
        ws.cell(row=k + 2, column=1, value=float(t_interp[k]))
        ws.cell(row=k + 2, column=2, value=float(interp_data['base_height'][k]))
        ws.cell(row=k + 2, column=3, value=float(interp_data['q_1'][k]))
        ws.cell(row=k + 2, column=4, value=float(interp_data['q_2'][k]))
        ws.cell(row=k + 2, column=5, value=float(interp_data['dz'][k]))
        ws.cell(row=k + 2, column=6, value=float(interp_data['dq_1'][k]))
        ws.cell(row=k + 2, column=7, value=float(interp_data['dq_2'][k]))
        ws.cell(row=k + 2, column=8, value=float(interp_data['tau_1'][k]))
        ws.cell(row=k + 2, column=9, value=float(interp_data['tau_2'][k]))
        ws.cell(row=k + 2, column=10, value=float(interp_data['tau_1_input'][k]))
        ws.cell(row=k + 2, column=11, value=float(interp_data['tau_2_input'][k]))
        ws.cell(row=k + 2, column=12, value=float(interp_data['grf_x'][k]))
        ws.cell(row=k + 2, column=13, value=float(interp_data['grf_z_spring'][k]))
        ws.cell(row=k + 2, column=14, value=float(interp_data['grf_z_body'][k]))
        ws.cell(row=k + 2, column=15, value=float(interp_data['hip_power'][k]))
        ws.cell(row=k + 2, column=16, value=float(interp_data['knee_power'][k]))
        ws.cell(row=k + 2, column=17, value=float(interp_data['total_power'][k]))
        ws.cell(row=k + 2, column=18, value=float(interp_data['hip_power_abs'][k]))
        ws.cell(row=k + 2, column=19, value=float(interp_data['knee_power_abs'][k]))
        ws.cell(row=k + 2, column=20, value=float(interp_data['total_power_abs'][k]))

    output_xlsx = "task0_vertjump_no_cvt_results.xlsx"
    wb.save(output_xlsx)
    print(f"\n[Data Save] Results saved to {output_xlsx} ({N_interp} rows, dt={dt_save}s)")

    np.savez('task0_vertjump_no_cvt_traj.npz',
             t=t_val, z=x_val[0,:], q1=x_val[1,:], q2=x_val[2,:],
             dz=v_val[0,:], dq1=v_val[1,:], dq2=v_val[2,:],
             tau1=u_tau_val[0,:], tau2=u_tau_val[1,:],
             grf_x=u_grf_val[0,:], grf_z=u_grf_val[1,:])
    print(f"[Diag] task0_vertjump_no_cvt_traj.npz saved (N={len(t_val)})")

    # Derived kinematics
    M_tot = M_tot_val
    A = params['m1']*params['r1'] + params['m_p']*params['r_p'] + params['m2']*params['l1']
    B = params['m2']*params['r2'] - params['m_c']*params['r_c'] - params['m_p']*params['l_c']

    z_com_start = (x_val[0,0]
                   + (A*np.sin(x_val[1,0]) + B*np.sin(x_val[1,0]+x_val[2,0])) / M_tot)
    z_com_end = (x_val[0,-1]
                 + (A*np.sin(x_val[1,-1]) + B*np.sin(x_val[1,-1]+x_val[2,-1])) / M_tot)
    v_com_z = (v_val[0,-1]
               + (A*np.cos(x_val[1,-1])*v_val[1,-1]
                  + B*np.cos(x_val[1,-1]+x_val[2,-1])*(v_val[1,-1]+v_val[2,-1])) / M_tot)

    q1_end, q2_end = x_val[1,-1], x_val[2,-1]
    c1_end, c2_end = np.cos(q1_end), np.cos(q2_end)
    c12_end = np.cos(q1_end + q2_end)
    K = params['m2']*params['l1']*params['r2'] - params['m_p']*params['l_c']*params['r_p']
    I_sig1 = ((params['I1'] + params['m1']*params['r1']**2)
              + (params['I_c'] + params['m_c']*params['r_c']**2)
              + (params['I_p'] + params['m_p']*params['r_p']**2 + params['m_p']*params['l_c']**2)
              + (params['I2'] + params['m2']*params['r2']**2 + params['m2']*params['l1']**2))
    I_sig2 = ((params['I2'] + params['m2']*params['r2']**2)
              + (params['I_c'] + params['m_c']*params['r_c']**2)
              + params['m_p']*params['l_c']**2)
    M_takeoff = np.array([
        [M_tot,                   A*c1_end + B*c12_end,      B*c12_end],
        [A*c1_end + B*c12_end,    I_sig1 + 2*K*c2_end,       I_sig2 + K*c2_end],
        [B*c12_end,               I_sig2 + K*c2_end,         I_sig2],
    ], dtype=float)
    v_takeoff = np.array([v_val[0,-1], v_val[1,-1], v_val[2,-1]], dtype=float)
    KE_total_takeoff = 0.5 * float(v_takeoff @ M_takeoff @ v_takeoff)
    KE_com_takeoff = 0.5 * M_tot * v_com_z**2
    KE_internal_takeoff = KE_total_takeoff - KE_com_takeoff
    delta_PE_base = M_tot*params['g']*(x_val[0,-1] - x_val[0,0])
    delta_PE_com_exact = M_tot*params['g']*(z_com_end - z_com_start)

    h_val_base     = x_val[0,-1] + v_val[0,-1]**2 / (2*params['g'])
    h_val_com      = z_com_end    + v_com_z**2      / (2*params['g'])
    h_val_base_com = x_val[0,-1] + v_com_z**2      / (2*params['g'])

    # Energy calculations for console output
    W_hip_cum = np.zeros(N); W_knee_cum = np.zeros(N); W_tot_cum = np.zeros(N)
    for k in range(1, N):
        ddt = t_val[k] - t_val[k-1]
        W_hip_cum[k]  = W_hip_cum[k-1]  + 0.5*(P_hip[k] +P_hip[k-1]  )*ddt
        W_knee_cum[k] = W_knee_cum[k-1] + 0.5*(P_knee[k]+P_knee[k-1] )*ddt
        W_tot_cum[k]  = W_tot_cum[k-1]  + 0.5*(P_total[k]+P_total[k-1])*ddt

    W_hip_total  = W_hip_cum[-1]
    W_knee_total = W_knee_cum[-1]
    W_mech_total = W_tot_cum[-1]

    _trap = np.trapezoid if hasattr(np, 'trapezoid') else np.trapz
    W_pos_hip  = _trap(np.maximum(P_hip,  0), t_val)
    W_neg_hip  = _trap(np.minimum(P_hip,  0), t_val)
    W_pos_knee = _trap(np.maximum(P_knee, 0), t_val)
    W_neg_knee = _trap(np.minimum(P_knee, 0), t_val)
    impulse_z  = _trap(u_grf_val[1,:], t_val)
    impulse_x  = _trap(u_grf_val[0,:], t_val)
    impulse_z_cum = np.array([_trap(u_grf_val[1,:k+1], t_val[:k+1]) for k in range(N)])
    impulse_x_cum = np.array([_trap(u_grf_val[0,:k+1], t_val[:k+1]) for k in range(N)])

    dq1_limit_arr = speed_torque_coeff*np.abs(u_tau_val[0,:]) + speed_torque_offset
    dq2_limit_arr = speed_torque_coeff*np.abs(u_tau_val[1,:]) + speed_torque_offset
    dq1_sat_arr = np.divide(np.abs(v_val[1,:]), dq1_limit_arr,
                            out=np.full_like(v_val[1,:], np.nan, dtype=float),
                            where=dq1_limit_arr > 1e-9)
    dq2_sat_arr = np.divide(np.abs(v_val[2,:]), dq2_limit_arr,
                            out=np.full_like(v_val[2,:], np.nan, dtype=float),
                            where=dq2_limit_arr > 1e-9)
    peak_dq1_idx = int(np.argmax(np.abs(v_val[1,:])))
    peak_dq2_idx = int(np.argmax(np.abs(v_val[2,:])))
    peak_sat1_idx = int(np.nanargmax(dq1_sat_arr))
    peak_sat2_idx = int(np.nanargmax(dq2_sat_arr))

    peak_events = [
        summarize_peak_event("Hip abs speed peak", peak_dq1_idx, t_val, v_val[1,:], u_tau_val[0,:],
                             u_grf_val[1,:], dq1_limit_arr, P_hip),
        summarize_peak_event("Knee abs speed peak", peak_dq2_idx, t_val, v_val[2,:], u_tau_val[1,:],
                             u_grf_val[1,:], dq2_limit_arr, P_knee),
        summarize_peak_event("Hip speed-limit saturation peak", peak_sat1_idx, t_val, v_val[1,:], u_tau_val[0,:],
                             u_grf_val[1,:], dq1_limit_arr, P_hip),
        summarize_peak_event("Knee speed-limit saturation peak", peak_sat2_idx, t_val, v_val[2,:], u_tau_val[1,:],
                             u_grf_val[1,:], dq2_limit_arr, P_knee),
    ]
    takeoff_event = {
        'time': float(t_val[-1]),
        'tau1': float(u_tau_val[0,-1]),
        'tau2': float(u_tau_val[1,-1]),
        'dq1': float(v_val[1,-1]),
        'dq2': float(v_val[2,-1]),
        'grf_z': float(u_grf_val[1,-1]),
        'dq1_limit': float(dq1_limit_arr[-1]),
        'dq2_limit': float(dq2_limit_arr[-1]),
        'dq1_sat_pct': float(100.0 * dq1_sat_arr[-1]),
        'dq2_sat_pct': float(100.0 * dq2_sat_arr[-1]),
        'hip_power': float(P_hip[-1]),
        'knee_power': float(P_knee[-1]),
    }

    v_z_from_grf  = (impulse_z - M_tot*params['g']*T_val) / M_tot
    h_grf         = v_z_from_grf**2 / (2*params['g'])
    h_grf_total   = x_val[0,0] + h_grf

    KE_from_mech = W_mech_total - delta_PE_base
    if KE_from_mech > 0:
        v_z_mech    = np.sqrt(2*KE_from_mech/M_tot)
        h_mech      = v_z_mech**2 / (2*params['g'])
        h_mech_total = x_val[0,-1] + h_mech
    else:
        v_z_mech = 0.0; h_mech = 0.0; h_mech_total = x_val[0,-1]

    energy_budget_residual = W_mech_total - delta_PE_com_exact - KE_total_takeoff

    if contact_model in ('hard', 'alpha'):
        fe = max(abs(x_val[0,k] + params['l1']*np.sin(x_val[1,k])
                     + params['l2']*np.sin(x_val[1,k]+x_val[2,k]))
                 for k in range(N))
    else:
        fe = max(delta_arr)   # soft: 최대 침투 깊이

    if contact_model in ('soft', 'soft_alpha'):
        E_spring_arr = 0.5 * params['k_c'] * delta_arr**2
        E_spring_max = E_spring_arr.max()
        E_spring_final = 0.5 * params['k_c'] * delta_arr[-1]**2
        E_damper = W_mech_total - KE_from_mech - delta_PE_base - E_spring_final
    else:
        E_spring_max = 0.0
        E_spring_final = 0.0
        E_damper = 0.0

    # === Console Output ===
    print(f"\n{'='*60}")
    print(f"  [{contact_model.upper()}] Jump Optimization Results")
    print(f"{'='*60}")
    print(f"  Jump Height (Base, raw dz)     : {h_val_base:.4f} m")
    print(f"  Jump Height (CoM)              : {h_val_com:.4f} m")
    print(f"  Jump Height (Base via CoM v_z) : {h_val_base_com:.4f} m  <- optimizer target")
    print(f"  Takeoff Base z        : {x_val[0,-1]:.4f} m")
    print(f"  Takeoff Base v_z      : {v_val[0,-1]:.4f} m/s")
    print(f"  Stance time           : {T_val:.4f} s")
    print(f"  GRF_z[-1]             : {u_grf_val[1,-1]:.4f} N")
    print(f"  GRF_z max             : {u_grf_val[1,:].max():.2f} N")
    if contact_model in ('soft', 'soft_alpha'):
        print(f"  Max foot penetration  : {fe*1000:.2f} mm")
        print(f"  Max spring energy     : {E_spring_max:.4f} J")
        print(f"  Damper dissipation    : {E_damper:.4f} J")

    # === 입력 에너지 계산 ===
    _trap = np.trapezoid if hasattr(np, 'trapezoid') else np.trapz
    E_fric_joint = _trap(JF_HIP * v_val[1,:]**2 + JF_KNEE * v_val[2,:]**2, t_val)
    E_fric_rail = _trap(RAIL_FRICTION * v_val[0,:]**2, t_val)
    W_input_total = W_mech_total + E_fric_joint + E_fric_rail

    # === Alpha 전후 Impulse ===
    impulse_z_spring = impulse_z  # GRF_spring (alpha 미적용)
    impulse_z_body = impulse_z * params.get('alpha', 1.0) if contact_model in ('alpha', 'soft_alpha') else impulse_z

    # === GRF alpha 전후 ===
    grf_z_spring = u_grf_val[1, :]  # raw foot-on-ground force
    # ALPHA FIX: grf_z_body = raw (alpha applied inside dynamics).
    grf_z_body = grf_z_spring

    print(f"\n{'─'*60}")
    print(f"  [Mechanical Energy (Output Torque)]")
    print(f"{'─'*60}")
    print(f"  W_hip  (total)        : {W_hip_total:+.4f} J  (+{W_pos_hip:.3f} / {W_neg_hip:.3f})")
    print(f"  W_knee (total)        : {W_knee_total:+.4f} J  (+{W_pos_knee:.3f} / {W_neg_knee:.3f})")
    print(f"  W_mech (output)       : {W_mech_total:+.4f} J")
    print(f"\n{'─'*60}")
    print(f"  [Input Energy (Output + Friction)]")
    print(f"{'─'*60}")
    print(f"  W_output (shaft)      : {W_mech_total:+.4f} J")
    print(f"  E_friction (joint)    : {E_fric_joint:+.4f} J")
    print(f"  E_friction (rail)     : {E_fric_rail:+.4f} J")
    print(f"  W_input (total)       : {W_input_total:+.4f} J")
    print(f"  v_z from mech energy  : {v_z_mech:.4f} m/s")
    print(f"  Jump h (mech energy)  : {h_mech_total:.4f} m")
    print(f"\n{'-'*60}")
    print(f"  [Energy Budget]")
    print(f"{'-'*60}")
    print(f"  COM z start           : {z_com_start:.4f} m")
    print(f"  COM z takeoff         : {z_com_end:.4f} m")
    print(f"  dPE base approx       : {delta_PE_base:.4f} J")
    print(f"  dPE COM exact         : {delta_PE_com_exact:.4f} J")
    print(f"  KE total @ takeoff    : {KE_total_takeoff:.4f} J")
    if contact_model in ('soft', 'soft_alpha'):
        print(f"  Spring E @ takeoff    : {E_spring_final:.4f} J")
    print(f"  Residual (W-dPE-KE)   : {energy_budget_residual:+.4f} J")
    print(f"  Residual note         : contact/spring-damper exchange + model residual")
    print(f"\n{'?'*60}")
    print(f"  [Takeoff Energy Split]")
    print(f"{'?'*60}")
    print(f"  KE_total @ takeoff    : {KE_total_takeoff:.4f} J")
    print(f"  KE_COM_z @ takeoff    : {KE_com_takeoff:.4f} J")
    print(f"  Internal KE residual  : {KE_internal_takeoff:.4f} J")
    print(f"  v_COM_z @ takeoff     : {v_com_z:.4f} m/s")
    if KE_total_takeoff > 1e-9:
        print(f"  COM share of KE       : {100.0*KE_com_takeoff/KE_total_takeoff:.2f} %")
        print(f"  Internal share of KE  : {100.0*KE_internal_takeoff/KE_total_takeoff:.2f} %")
    print(f"\n{'─'*60}")
    print(f"  [GRF Impulse]       ")
    print(f"{'─'*60}")
    print(f"  Impulse_z (spring)    : {impulse_z_spring:.4f} N·s  (alpha 미적용, 로드셀 기준)")
    print(f"  Impulse_z (body)      : {impulse_z_body:.4f} N·s  (alpha 적용, body가 받는 impulse)")
    print(f"  v_z from GRF impulse  : {v_z_from_grf:.4f} m/s")
    print(f"  Jump h (GRF impulse)  : {h_grf_total:.4f} m")
    print(f"\n{'?'*60}")
    print(f"  [Peak Joint-Speed Diagnostics]")
    print(f"{'?'*60}")
    for ev in peak_events:
        print(f"  {ev['name']}")
        print(f"    t={ev['time']:.4f} s, dq={ev['dq']:+.4f} rad/s, tau={ev['tau']:+.4f} Nm")
        print(f"    GRF_z={ev['grf_z']:+.4f} N, power={ev['power']:+.4f} W")
        print(f"    motor speed limit={ev['speed_limit']:.4f} rad/s, saturation={ev['sat_pct']:.2f} %")
    print(f"\n{'?'*60}")
    print(f"  [Takeoff Motor State]")
    print(f"{'?'*60}")
    print(f"  t_takeoff            : {takeoff_event['time']:.4f} s")
    print(f"  GRF_z @ takeoff      : {takeoff_event['grf_z']:+.4f} N")
    print(f"  Hip  tau/dq/lim      : {takeoff_event['tau1']:+.4f} Nm / {takeoff_event['dq1']:+.4f} rad/s / {takeoff_event['dq1_limit']:.4f} rad/s")
    print(f"  Hip  sat/power       : {takeoff_event['dq1_sat_pct']:.2f} % / {takeoff_event['hip_power']:+.4f} W")
    print(f"  Knee tau/dq/lim      : {takeoff_event['tau2']:+.4f} Nm / {takeoff_event['dq2']:+.4f} rad/s / {takeoff_event['dq2_limit']:.4f} rad/s")
    print(f"  Knee sat/power       : {takeoff_event['dq2_sat_pct']:.2f} % / {takeoff_event['knee_power']:+.4f} W")
    print(f"{'='*60}\n")

    # === Figure 1: Kinematics & Dynamics ===
    model_label = contact_model.upper()
    if contact_model in ('soft', 'soft_alpha'):
        model_label += f" (k_c={k_c:.0f}, b_c={b_c:.0f})"

    plt.figure(figsize=(15, 8))
    plt.suptitle(f"[{model_label}] Kinematics & Dynamics", fontsize=12)

    plt.subplot(3, 2, 1)
    plt.plot(t_val, x_val[0,:], lw=2)
    plt.title("Base Height (m)"); plt.grid(True)

    plt.subplot(3, 2, 2)
    plt.plot(t_val, v_val[0,:], lw=2)
    plt.title("Base Vertical Velocity (m/s)"); plt.grid(True)

    plt.subplot(3, 2, 3)
    plt.plot(t_val, np.degrees(x_val[1,:]), label='q1 (Hip)')
    plt.plot(t_val, np.degrees(x_val[2,:]), label='q2 (Knee)')
    plt.title("Joint Angles (deg)"); plt.legend(); plt.grid(True)

    plt.subplot(3, 2, 4)
    plt.plot(t_val, v_val[1,:], label='dq1')
    plt.plot(t_val, v_val[2,:], label='dq2')
    plt.hlines([-dq_lim, dq_lim], 0, t_val[-1], 'k', '--')
    plt.title("Joint Velocities (rad/s)"); plt.legend(); plt.grid(True)

    plt.subplot(3, 2, 5)
    plt.plot(t_val, u_tau_val[0,:], label='Tau1 (Hip)')
    plt.plot(t_val, u_tau_val[1,:], label='Tau2 (Knee)')
    plt.hlines([-tau_lim, tau_lim], 0, t_val[-1], 'k', '--')
    plt.title("Joint Torques (Nm)"); plt.legend(); plt.grid(True)

    plt.subplot(3, 2, 6)
    for j in range(2):
        tr = np.linspace(-tau_lim, tau_lim, 50)
        plt.plot(tr, speed_torque_coeff*np.abs(tr)+speed_torque_offset, 'k--', alpha=0.3)
        plt.plot(u_tau_val[j,:], np.abs(v_val[j+1,:]), '.', label=f'Motor {j+1}')
    plt.title("T-N Limit Check"); plt.legend(); plt.grid(True)
    plt.tight_layout()

    # === Figure 2: Energy & GRF Analysis ===
    fig4 = plt.figure(figsize=(14, 10))
    fig4.suptitle(
        f"[{model_label}] Mechanical Energy Analysis\n"
        f"W_hip={W_hip_total:+.3f}J  W_knee={W_knee_total:+.3f}J  "
        f"W_mech={W_mech_total:+.3f}J  |  "
        f"Impulse_z={impulse_z:.3f}N·s  h={h_val_base_com:.3f}m",
        fontsize=11)
    gs4 = GridSpec(2, 2, figure=fig4, hspace=0.4, wspace=0.35)

    ax4a = fig4.add_subplot(gs4[0,0])
    ax4a.plot(t_val, P_hip,  '#2971B1', lw=2, label='Hip power')
    ax4a.plot(t_val, P_knee,  '#C0392B', lw=2, label='Knee power')
    ax4a.plot(t_val, P_total, 'k',       lw=2, ls='--', label='Total')
    ax4a.fill_between(t_val, P_total, 0, where=(P_total>=0), alpha=0.15, color='green')
    ax4a.fill_between(t_val, P_total, 0, where=(P_total<0),  alpha=0.15, color='red')
    ax4a.axhline(0, color='k', lw=0.8)
    ax4a.set(xlabel='Time (s)', ylabel='Power (W)', title='Instantaneous Power')
    ax4a.legend(fontsize=8); ax4a.grid(True, alpha=0.3)

    ax4b = fig4.add_subplot(gs4[0,1])
    ax4b.plot(t_val, W_hip_cum,  '#2971B1', lw=2, label=f'W_hip  ({W_hip_total:+.2f}J)')
    ax4b.plot(t_val, W_knee_cum, '#C0392B', lw=2, label=f'W_knee ({W_knee_total:+.2f}J)')
    ax4b.plot(t_val, W_tot_cum,  'k',       lw=2.5, ls='--',
              label=f'W_total ({W_mech_total:+.2f}J)')
    ax4b.axhline(36.913, color='gray', ls=':', lw=1.2, label='Real P40 total: 36.9J')
    ax4b.axhline(0, color='k', lw=0.8)
    ax4b.set(xlabel='Time (s)', ylabel='Cumul. work (J)', title='Cumulative Energy')
    ax4b.legend(fontsize=7); ax4b.grid(True, alpha=0.3)

    ax4c = fig4.add_subplot(gs4[1,0])
    ax4c2 = ax4c.twinx()
    ax4c.plot(t_val, u_grf_val[1,:], '#2E86AB', lw=2, label='GRF_z: ground→robot (N)')
    ax4c.plot(t_val, u_grf_val[0,:], '#E84855', lw=2, label='GRF_x (N)')
    ax4c2.plot(t_val, impulse_z_cum, '#2E86AB', lw=1.5, ls='--', alpha=0.7,
               label=f'Impulse_z ({impulse_z:.3f} N·s)')
    ax4c2.axhline(20.35, color='gray', ls=':', lw=1.2, label='Real P40: 20.35 N·s')
    ax4c.set(xlabel='Time (s)', ylabel='Force (N)', title='GRF: ground→robot')
    ax4c2.set_ylabel('Cumul. impulse (N·s)')
    l1_, la1 = ax4c.get_legend_handles_labels()
    l2_, la2 = ax4c2.get_legend_handles_labels()
    ax4c.legend(l1_+l2_, la1+la2, fontsize=7, loc='lower center')
    ax4c.grid(True, alpha=0.3)

    ax4d = fig4.add_subplot(gs4[1,1])
    methods = ['Base\n(raw dz)', 'Base\n(via CoM v_z)', 'CoM',
               'GRF\nImpulse', 'Mech.\nEnergy']
    heights = [h_val_base, h_val_base_com, h_val_com, h_grf_total, h_mech_total]
    cols    = ['#95A5A6','#5B9BD5','#76A5E3','#C0392B','#E67E22']
    bars = ax4d.bar(methods, heights, color=cols, alpha=0.8, edgecolor='k', lw=0.8)
    for bar, h in zip(bars, heights):
        ax4d.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.005,
                  f'{h:.3f}m', ha='center', va='bottom', fontsize=9, fontweight='bold')
    ax4d.axhline(0.861, color='gray', ls=':', lw=1.2, label='Real P40: 0.861m')
    ax4d.set(ylabel='Jump height (m)', title='Jump Height Comparison')
    ax4d.legend(fontsize=8); ax4d.grid(True, alpha=0.3, axis='y')
    ax4d.set_ylim(0, max(heights)*1.18)

    # === Figure 3: Stick Figure ===
    fig_stick, ax_stick = plt.subplots(figsize=(8, 8))
    ax_stick.plot([-0.5, 0.5], [0, 0], 'k-', lw=3)
    if contact_model in ('soft', 'soft_alpha'):
        ax_stick.axhline(-delta_arr.max()*10, color='brown', ls=':', alpha=0.5,
                    label=f'Max penetration (×10): {delta_arr.max()*1000:.1f}mm')
    indices = np.linspace(0, N-1, 5, dtype=int)
    colors  = plt.cm.viridis(np.linspace(0, 1, len(indices)))
    l1, l2 = params['l1'], params['l2']
    for idx, i in enumerate(indices):
        zi, q1i, q2i = x_val[0,i], x_val[1,i], x_val[2,i]
        hx, hy = 0, zi
        kx = hx + l1*np.cos(q1i); ky = hy + l1*np.sin(q1i)
        fx = kx + l2*np.cos(q1i+q2i); fy = ky + l2*np.sin(q1i+q2i)
        ax_stick.plot([hx,kx,fx],[hy,ky,fy],'o-',lw=2,color=colors[idx],
                      label=f't={t_val[i]:.2f}s')
    ax_stick.set_title(f"[{model_label}] Stick Figure (H={h_val_base_com:.3f}m)")
    ax_stick.set_xlim([-0.1, 0.3])
    ax_stick.set_ylim([-0.05, 0.55])
    ax_stick.set_aspect('equal', adjustable='box')
    ax_stick.legend()
    ax_stick.grid(True)

    # === Figure 4: Input vs Output Torque ===
    fig4, axes4 = plt.subplots(2, 1, figsize=(12, 8))
    fig4.suptitle(f"[{model_label}] Input vs Output Torque", fontsize=12)

    # 입력 토크 = 출력 토크 + 마찰 토크
    tau1_input = u_tau_val[0,:] + JF_HIP * v_val[1,:]
    tau2_input = u_tau_val[1,:] + JF_KNEE * v_val[2,:]

    axes4[0].plot(t_val, u_tau_val[0,:], 'b-', lw=2, label='Hip output (shaft)')
    axes4[0].plot(t_val, tau1_input, 'b--', lw=1.5, label='Hip input (shaft+friction)')
    axes4[0].plot(t_val, u_tau_val[1,:], 'r-', lw=2, label='Knee output (shaft)')
    axes4[0].plot(t_val, tau2_input, 'r--', lw=1.5, label='Knee input (shaft+friction)')
    axes4[0].axhline(0, color='k', lw=0.5)
    axes4[0].set(title='Torque: Output vs Input', ylabel='Torque (Nm)', xlabel='Time (s)')
    axes4[0].legend(fontsize=8); axes4[0].grid(True, alpha=0.3)

    axes4[1].plot(t_val, JF_HIP * v_val[1,:], 'b-', lw=1.5, label='Hip friction torque')
    axes4[1].plot(t_val, JF_KNEE * v_val[2,:], 'r-', lw=1.5, label='Knee friction torque')
    axes4[1].plot(t_val, RAIL_FRICTION * v_val[0,:], 'k--', lw=1.5, label='Rail friction force')
    axes4[1].axhline(0, color='k', lw=0.5)
    axes4[1].set(title='Friction Torque/Force', ylabel='Nm or N', xlabel='Time (s)')
    axes4[1].legend(fontsize=8); axes4[1].grid(True, alpha=0.3)
    fig4.tight_layout()

    # === Figure 5: Input vs Output Energy ===
    fig5, axes5 = plt.subplots(2, 1, figsize=(12, 8))
    fig5.suptitle(f"[{model_label}] Input vs Output Energy", fontsize=12)

    P_output = P_hip + P_knee
    P_fric = JF_HIP * v_val[1,:]**2 + JF_KNEE * v_val[2,:]**2 + RAIL_FRICTION * v_val[0,:]**2
    P_input = P_output + P_fric

    W_out_cum = np.zeros(N); W_in_cum = np.zeros(N); W_fric_cum = np.zeros(N)
    for k in range(1, N):
        ddt = t_val[k] - t_val[k-1]
        W_out_cum[k] = W_out_cum[k-1] + 0.5*(P_output[k]+P_output[k-1])*ddt
        W_fric_cum[k] = W_fric_cum[k-1] + 0.5*(P_fric[k]+P_fric[k-1])*ddt
        W_in_cum[k] = W_out_cum[k] + W_fric_cum[k]

    axes5[0].plot(t_val, P_output, 'b-', lw=2, label='Output power (shaft)')
    axes5[0].plot(t_val, P_input, 'r--', lw=1.5, label='Input power (shaft+friction)')
    axes5[0].plot(t_val, P_fric, 'k:', lw=1, label='Friction power')
    axes5[0].axhline(0, color='k', lw=0.5)
    axes5[0].set(title='Instantaneous Power', ylabel='Power (W)', xlabel='Time (s)')
    axes5[0].legend(fontsize=8); axes5[0].grid(True, alpha=0.3)

    axes5[1].plot(t_val, W_out_cum, 'b-', lw=2, label=f'Output energy ({W_mech_total:.1f} J)')
    axes5[1].plot(t_val, W_in_cum, 'r--', lw=2, label=f'Input energy ({W_input_total:.1f} J)')
    axes5[1].plot(t_val, W_fric_cum, 'k:', lw=1.5, label=f'Friction loss ({E_fric_joint+E_fric_rail:.1f} J)')
    axes5[1].set(title='Cumulative Energy', ylabel='Energy (J)', xlabel='Time (s)')
    axes5[1].legend(fontsize=8); axes5[1].grid(True, alpha=0.3)
    fig5.tight_layout()

    # === Figure 6: GRF before/after Alpha ===
    if contact_model in ('alpha', 'soft_alpha'):
        fig6, axes6 = plt.subplots(2, 1, figsize=(12, 8))
        fig6.suptitle(f"[{model_label}] GRF: Before vs After Alpha ({alpha:.2f})", fontsize=12)

        axes6[0].plot(t_val, grf_z_spring, 'k-', lw=2, label=f'GRF spring (loadcell, Imp={impulse_z_spring:.2f})')
        axes6[0].plot(t_val, grf_z_body, 'r--', lw=2, label=f'GRF body (alpha*spring, Imp={impulse_z_body:.2f})')
        axes6[0].fill_between(t_val, grf_z_spring, grf_z_body, alpha=0.15, color='red',
                              label=f'Absorbed ({(1-alpha)*100:.0f}%)')
        axes6[0].axhline(0, color='k', lw=0.5)
        axes6[0].set(title='GRF: Spring vs Body', ylabel='GRF_z (N)', xlabel='Time (s)')
        axes6[0].legend(fontsize=8); axes6[0].grid(True, alpha=0.3)

        imp_spring_cum = np.array([_trap(grf_z_spring[:k+1], t_val[:k+1]) for k in range(N)])
        imp_body_cum = np.array([_trap(grf_z_body[:k+1], t_val[:k+1]) for k in range(N)])
        axes6[1].plot(t_val, imp_spring_cum, 'k-', lw=2, label=f'Impulse spring ({impulse_z_spring:.2f})')
        axes6[1].plot(t_val, imp_body_cum, 'r--', lw=2, label=f'Impulse body ({impulse_z_body:.2f})')
        axes6[1].set(title='Cumulative Impulse', ylabel='Impulse (N*s)', xlabel='Time (s)')
        axes6[1].legend(fontsize=8); axes6[1].grid(True, alpha=0.3)
        fig6.tight_layout()

    # === Figure 7: Motor mechanical energy vs internal (link rotation) KE ===
    A_cad = params['m1']*params['r1'] + params['m_p']*params['r_p'] + params['m2']*params['l1']
    B_cad = params['m2']*params['r2'] - params['m_c']*params['r_c'] - params['m_p']*params['l_c']
    K_cad = params['m2']*params['l1']*params['r2'] - params['m_p']*params['l_c']*params['r_p']
    I_sig1_cad = ((params['I1'] + params['m1']*params['r1']**2)
                  + (params['I_c'] + params['m_c']*params['r_c']**2)
                  + (params['I_p'] + params['m_p']*params['r_p']**2 + params['m_p']*params['l_c']**2)
                  + (params['I2'] + params['m2']*params['r2']**2 + params['m2']*params['l1']**2))
    I_sig2_cad = ((params['I2'] + params['m2']*params['r2']**2)
                  + (params['I_c'] + params['m_c']*params['r_c']**2)
                  + params['m_p']*params['l_c']**2)
    KE_total_t = np.zeros(N); KE_com_t = np.zeros(N); v_com_z_t = np.zeros(N)
    for k in range(N):
        c1 = np.cos(x_val[1,k]); c2 = np.cos(x_val[2,k]); c12 = np.cos(x_val[1,k]+x_val[2,k])
        Mk = np.array([
            [M_tot,                A_cad*c1+B_cad*c12,        B_cad*c12             ],
            [A_cad*c1+B_cad*c12,   I_sig1_cad+2*K_cad*c2,     I_sig2_cad+K_cad*c2   ],
            [B_cad*c12,            I_sig2_cad+K_cad*c2,       I_sig2_cad            ],
        ])
        vk = v_val[:,k]
        KE_total_t[k] = 0.5*float(vk @ Mk @ vk)
        v_com_z_t[k]  = (v_val[0,k] + (A_cad*c1*v_val[1,k]
                          + B_cad*c12*(v_val[1,k]+v_val[2,k]))/M_tot)
        KE_com_t[k]   = 0.5*M_tot*v_com_z_t[k]**2
    KE_internal_t = KE_total_t - KE_com_t

    fig7, ax7 = plt.subplots(figsize=(12, 6))
    fig7.suptitle(f"[{model_label}] Motor mechanical energy vs internal (link rotation) KE",
                  fontsize=12)
    ax7.plot(t_val*1000, W_tot_cum,     'k-',  lw=2.5,
             label=f'W_mech (cumul motor output) = {W_mech_total:.2f} J')
    ax7.plot(t_val*1000, KE_internal_t, 'r-',  lw=2,
             label=f'KE_internal (link rotation) — final {KE_internal_t[-1]:.2f} J')
    ax7.plot(t_val*1000, KE_com_t,      'b-',  lw=2,
             label=f'KE_COM_z (body translation) — final {KE_com_t[-1]:.2f} J')
    ax7.plot(t_val*1000, KE_total_t,    'g--', lw=1.5, alpha=0.7,
             label=f'KE_total = KE_int + KE_COM — final {KE_total_t[-1]:.2f} J')
    ax7.axhline(0, color='gray', lw=0.5)
    ax7.set(xlabel='time (ms)', ylabel='Energy (J)')
    ax7.set_title(f'Internal share at takeoff: {100*KE_internal_t[-1]/max(KE_total_t[-1],1e-9):.1f}%')
    ax7.legend(fontsize=10, loc='best'); ax7.grid(alpha=0.3)
    fig7.tight_layout()

    pass  # plt.show() disabled

    # === Animation generation (Task animation hook) ===
    try:
        from animate_results import animate_jump
        animate_jump(output_xlsx)
    except Exception as _e:
        print(f'[anim] skipped: {_e}')

    # === PLOT_HELPER_BLOCK ===
    try:
        from task_plot_helper import regen_one_task
        regen_one_task(output_xlsx, out_dir='plots/regen')
        print(f'[plots] regen done: {output_xlsx}')
    except Exception as _e:
        print(f'[plots] skipped: {_e}')


if __name__ == "__main__":
    optimize_jumping()
